"""Test: openpyxl"""
import sys
sys.stdout.reconfigure(line_buffering=True)
try:
    from openpyxl import Workbook

    # Create a workbook and write data
    wb = Workbook()
    ws = wb.active
    ws.title = "Test"
    ws["A1"] = 42
    ws["B1"] = "hello"
    ws.append([1, 2, 3])

    # Verify data
    assert ws["A1"].value == 42, f"A1 expected 42, got {ws['A1'].value}"
    assert ws["B1"].value == "hello", f"B1 expected 'hello', got {ws['B1'].value}"
    assert ws.title == "Test", f"title expected 'Test', got {ws.title}"

    print("openpyxl: PASS")
except Exception as e:
    print(f"openpyxl: FAIL: {e}")
    sys.exit(1)
